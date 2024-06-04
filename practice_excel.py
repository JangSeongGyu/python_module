from openpyxl import Workbook,load_workbook
# wb = Workbook()

wb = Workbook()
ws= wb.create_sheet("TestSheet")


count = 0
sum=0
for row in ws["B2:E5"]:
    for cell in row:
        count+=1
        cell.value=count
        sum+=count
        
ws["B8"] = "平均"
ws["B9"] = sum/count
ws["C8"] = "合計"
ws["C9"] = sum

input_ws= wb.create_sheet("DataInput")


for i in range(1,11):
    input_ws.cell(i,1,i)

for i in range(1,11):
    cell_data = input_ws.cell(i,1).value
    ws.cell(i+1, 6, cell_data+10)
    
    
wb.save("test.xlsx")